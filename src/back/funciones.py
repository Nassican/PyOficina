import openpyxl

archivo_excel = "./src/inventario.xlsx"

def listar_categorias(archivo_excel):
    try:
        libro = openpyxl.load_workbook(archivo_excel)
        categorias = libro.sheetnames
        print("Las categorias son: ", categorias)
        libro.close()
        return categorias
    except Exception as e:
        print(f"Error al cargar el archivo: {e}")
        return []

def listar_productos(categoria, archivo_excel):
    try:
        libro = openpyxl.load_workbook(archivo_excel)
        if categoria not in libro.sheetnames:
            print("Categoria no encontrada")
            return []

        hoja = libro[categoria]
        productos = []
        for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row, max_col=hoja.max_column):
            for celda in fila:
                if celda.value == "DETALLE":
                    columna_detalle = celda.column
                    for producto_celda in hoja.iter_rows(min_row=celda.row + 1, max_col=columna_detalle, max_row=hoja.max_row):
                        producto = producto_celda[columna_detalle - 1].value
                        if producto:
                            productos.append(producto)
                    return productos
        print("No se encontraron productos en la categoria")
        libro.close()
        return productos
    except Exception as e:
        print(f"Error al listar productos: {e}")
        return []

def buscar_codigo_producto(categoria, nombre_producto, archivo_excel, columnas_codigo=1):
    try:
        libro = openpyxl.load_workbook(archivo_excel)
        if categoria not in libro.sheetnames:
            print("Categoria no encontrada")
            libro.close()
            return None

        hoja = libro[categoria]
        for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row, max_col=hoja.max_column):
            for celda in fila:
                if celda.value == "DETALLE":
                    columna_detalle = celda.column
                    for producto_celda in hoja.iter_rows(min_row=celda.row + 1, max_col=columna_detalle, max_row=hoja.max_row):
                        if producto_celda[columna_detalle - 1].value == nombre_producto:
                            codigo_celda = producto_celda[columna_detalle - 1].offset(column=columnas_codigo)
                            columna_codigo = codigo_celda.column
                            fila_numero_codigo = codigo_celda.row
                            columna_letra_codigo = openpyxl.utils.get_column_letter(columna_codigo)

                            # Cogemos la fila en donde esta el producto
                            fila_numero_producto = producto_celda[columna_detalle - 1].row
                            columna_letra_producto = openpyxl.utils.get_column_letter(columna_detalle).upper()
                            print(f"El código del producto '{nombre_producto}' es: {codigo_celda.value} (celda {columna_letra_codigo}{fila_numero_codigo})")
                            return codigo_celda.value, columna_letra_codigo, fila_numero_codigo, columna_letra_producto, fila_numero_producto
        print(f"No se encontró el producto '{nombre_producto}' en la categoría '{categoria}'")
        return None
    except Exception as e:
        print(f"Error al buscar código de producto: {e}")
        return None
    finally:
        if libro:
            libro.close()

def guardar_edicion_producto(archivo_excel, nombre_producto, codigo_producto, categoria, fila_numero_producto, columna_letra_producto, fila_numero_codigo, columna_letra_codigo):
    try:
        libro = openpyxl.load_workbook(archivo_excel)
        hoja = libro[categoria]  # Seleccionar la hoja de la categoría actual

        # Actualizar el nombre del producto en la celda correspondiente
        hoja[f"{columna_letra_producto}{fila_numero_producto}"] = nombre_producto

        # Actualizar el código del producto en la celda correspondiente
        hoja[f"{columna_letra_codigo}{fila_numero_codigo}"] = int(codigo_producto)

        # Guardar los cambios en el archivo Excel
        libro.save(archivo_excel)
        libro.close()
    except Exception as e:
        print(f"Error al guardar la edición del producto: {e}")
        raise e

def agregar_producto(archivo_excel, categoria, nombre_producto, codigo_producto, columnas_codigo=1):
    libro = None
    try:
        libro = openpyxl.load_workbook(archivo_excel)
        hoja = libro[categoria]  # Seleccionar la hoja de la categoría actual

        for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row, max_col=hoja.max_column):
            for celda in fila:
                if celda.value == "DETALLE":
                    columna_detalle = celda.column
                    for producto_celda in hoja.iter_rows(min_row=celda.row + 1, max_col=columna_detalle, max_row=hoja.max_row):
                        ultima_fila = hoja.max_row
                        hoja.cell(row=ultima_fila + 1, column=columna_detalle).value = nombre_producto
                        hoja.cell(row=ultima_fila + 1, column=columna_detalle + 1).value = codigo_producto

                        # Guardar los cambios en el archivo Excel
                        print(f"Se ha agregado el producto '{nombre_producto}' con código '{codigo_producto}' a la categoría '{categoria}'")
                        libro.save(archivo_excel)
                        return  # Salir después de agregar el producto
        print("No se encontró la columna 'DETALLE' en la hoja.")
    except Exception as e:
        print(f"Error al agregar el producto: {e}")
    finally:
        if libro:
            libro.close()


# Funcion para obtener el total de productos y guardarlos en un array de todas las categorias
def conseguir_total_productos(archivo_excel):
    try:
        libro = openpyxl.load_workbook(archivo_excel)
        categorias = libro.sheetnames
        productos = []
        for categoria in categorias:
            hoja = libro[categoria]
            for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row, max_col=hoja.max_column):
                for celda in fila:
                    if celda.value == "DETALLE":
                        columna_detalle = celda.column
                        for producto_celda in hoja.iter_rows(min_row=celda.row + 1, max_col=columna_detalle, max_row=hoja.max_row):
                            producto = producto_celda[columna_detalle - 1].value
                            if producto:
                                productos.append(producto)
        libro.close()
        return productos
    except Exception as e:
        print(f"Error al cargar el archivo: {e}")
        return []


def while_loop(archivo_excel):
    while True:
        categoria = input("Digite la categoria del producto: ")
        nombre_producto = input("Digite el nombre del producto: ")
        columnas_codigo = int(input("Ingrese el número de columnas a la derecha de 'DETALLE' donde se encuentra el código del producto: "))

        codigo_producto = buscar_codigo_producto(categoria, nombre_producto, archivo_excel, columnas_codigo)

        if codigo_producto is not None:
            print(f"El código del producto '{nombre_producto}' es: {codigo_producto}")
        else:
            print(f"No se encontró el producto '{nombre_producto}' en el archivo.")

        resp = input("¿Desea continuar? (si/no): ").strip().lower()
        if resp != "si":
            break

if __name__ == '__main__':
    listar_categorias(archivo_excel)
    while_loop(archivo_excel)
    print("Fin del programa")
